import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def fill_excel_template(excel_template_file, output_file, data_to_insert):
    try:
        wb = load_workbook(filename=excel_template_file)
        ws = wb["Sheet1"]
        for cell_id, value in data_to_insert.items():
            ws[cell_id] = value
        wb.save(output_file)
        return True, None
    except FileNotFoundError:
        return False, "Template file not found."
    except InvalidFileException:
        return False, "Invalid Excel file."
    except KeyError:
        return False, "Sheet1 not found in the workbook."
    except Exception as e:
        return False, str(e)