from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
import os

def convert_excel_to_markdown(excel_file_path):
    try:
        wb = load_workbook(excel_file_path, data_only=True)
        markdown_output = []
        
        # Add the filename as the main header
        filename = os.path.basename(excel_file_path)
        markdown_output.append(f"# {filename}\n")

        for ws in wb.worksheets:
            # Add worksheet title as header
            markdown_output.append(f"\n## {ws.title}\n")
            
            # Create a dictionary to store merged cell ranges
            merged_ranges = {}
            for merged_range in ws.merged_cells.ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell_id = f"{get_column_letter(col)}{row}"
                        merged_ranges[cell_id] = merged_range.coord
            
            # Iterate through all cells in the worksheet
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    cell_id = f"{get_column_letter(col_idx)}{row_idx}"
                    cell_value = cell.value
                    
                    # Only include non-empty cells
                    if cell_value is not None and str(cell_value).strip() != "":
                        # Check if cell is part of a merged range
                        merge_info = ""
                        if cell_id in merged_ranges:
                            merge_info = f" (merged range: {merged_ranges[cell_id]})"
                        
                        markdown_output.append(f"{cell_id}: \"{str(cell_value)}\"{merge_info}  ")

        return True, '\n'.join(markdown_output)
    except FileNotFoundError:
        return False, "Excel file not found."
    except InvalidFileException:
        return False, "Invalid Excel file."
    except Exception as e:
        return False, str(e) 