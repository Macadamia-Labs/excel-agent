import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import os # Added for basename

def fill_excel_template(excel_template_file, output_file, data_to_insert):
    print(f"Starting to fill Excel template '{os.path.basename(excel_template_file)}'...")
    try:
        wb = load_workbook(filename=excel_template_file)
        ws = wb["Sheet1"]
        
        # Create a dictionary to store merged cell ranges
        merged_ranges = {}
        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_id = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    merged_ranges[cell_id] = merged_range.coord
        
        for cell_id, value in data_to_insert.items():
            try:
                # Check if the cell is part of a merged range
                if cell_id in merged_ranges:
                    # Get the master cell (top-left cell) of the merged range
                    master_cell = ws[merged_ranges[cell_id].split(':')[0]]
                    master_cell.value = value
                else:
                    ws[cell_id] = value
            except Exception as cell_error:
                print(f"Error setting cell {cell_id} to {value}: {cell_error}")
                continue
                
        wb.save(output_file)
        print(f"Successfully filled Excel template and saved to '{os.path.basename(output_file)}'.")
        return True, None
    except FileNotFoundError:
        print(f"Error filling template '{os.path.basename(excel_template_file)}': Template file not found.")
        return False, "Template file not found."
    except InvalidFileException:
        print(f"Error filling template '{os.path.basename(excel_template_file)}': Invalid Excel file.")
        return False, "Invalid Excel file."
    except KeyError:
        print(f"Error filling template '{os.path.basename(excel_template_file)}': Sheet1 not found.")
        return False, "Sheet1 not found in the workbook."
    except Exception as e:
        print(f"Error filling template '{os.path.basename(excel_template_file)}': {str(e)}")
        return False, str(e)