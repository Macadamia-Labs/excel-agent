from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
import os

def convert_excel_to_markdown(excel_file_path: str) -> tuple[bool, str]:
    """
    Converts an Excel file to a markdown string.

    Args:
        excel_file_path: Path to the input Excel file.

    Returns:
        A tuple containing:
        - bool: True if conversion was successful, False otherwise.
        - str: The generated markdown content or an error message.
    """
    filename_for_log = os.path.basename(excel_file_path)
    print(f"Starting conversion of '{filename_for_log}' to Markdown...")
    try:
        wb = load_workbook(excel_file_path, data_only=True) # data_only returns the *value* of a formula
        markdown_output = []
        
        # Add the filename as the main header
        filename = os.path.basename(excel_file_path)
        markdown_output.append(f"# {filename}\n")

        for ws in wb.worksheets:
            # Add worksheet title as header
            markdown_output.append(f"\n## {ws.title}\n")
            
            # Create a dictionary to store merged cell ranges
            merged_ranges = {}
            for merged_range in ws.merged_cells.ranges:
                # Ensure the range coordinates are valid before iterating
                if merged_range.min_row is None or merged_range.max_row is None or \
                   merged_range.min_col is None or merged_range.max_col is None:
                    continue # Skip invalid ranges if any
                    
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        # Handle potential issues with get_column_letter if col is large or 0
                        try:
                            col_letter = get_column_letter(col)
                        except ValueError:
                            continue # Skip if column index is invalid
                        cell_id = f"{col_letter}{row}"
                        merged_ranges[cell_id] = merged_range.coord
            
            # Iterate through all cells in the worksheet
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    try:
                        col_letter = get_column_letter(col_idx)
                    except ValueError:
                        continue # Skip invalid column index
                    cell_id = f"{col_letter}{row_idx}"
                    cell_value = cell.value
                    
                    # Only include non-empty cells
                    if cell_value is not None and str(cell_value).strip() != "":
                        # Check if cell is part of a merged range
                        merge_info = ""
                        if cell_id in merged_ranges:
                            # Ensure merged_ranges[cell_id] is not None before formatting
                            coord = merged_ranges.get(cell_id)
                            if coord:
                                merge_info = f" (merged range: {coord})"
                        
                        # Represent the cell value as a string, handle potential errors
                        try:
                            value_str = str(cell_value)
                        except Exception:
                            value_str = "[Error converting value]"
                            
                        markdown_output.append(f"{cell_id}: \"{value_str}\"{merge_info}  ")

        print(f"Successfully converted '{filename_for_log}' to Markdown.")
        return True, '\n'.join(markdown_output)
        
    except FileNotFoundError:
        print(f"Error converting '{filename_for_log}': File not found.")
        return False, f"Excel file not found at path: {excel_file_path}"
    except InvalidFileException:
        print(f"Error converting '{filename_for_log}': Invalid or corrupted file.")
        return False, f"Invalid or corrupted Excel file: {excel_file_path}"
    except Exception as e:
        # Catch other potential exceptions during loading or processing
        print(f"Error converting '{filename_for_log}': {str(e)}")
        return False, f"An unexpected error occurred processing {os.path.basename(excel_file_path)}: {str(e)}" 