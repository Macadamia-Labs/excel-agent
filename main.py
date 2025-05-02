from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel_to_markdown(filename):
    wb = load_workbook(f"input/{filename}", data_only=True)  # data_only returns the *value* of a formula
    markdown_output = []
    
    # Add the filename as the main header
    markdown_output.append(f"# {filename}\n")

    for ws in wb.worksheets:
        # Add worksheet title as header
        markdown_output.append(f"\n## {ws.title}\n")
        
        # Create a dictionary to store merged cell ranges
        merged_ranges = {}
        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_id = f"{get_column_letter(col)}{row}"
                    merged_ranges[cell_id] = merged_range.coord
        
        # Iterate through all cells in the worksheet
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_id = f"{get_column_letter(col_idx)}{row_idx}"
                cell_value = cell.value
                
                # Only include non-empty cells
                if cell_value is not None and str(cell_value).strip() != "":
                    # Check if cell is part of a merged range
                    merge_info = ""
                    if cell_id in merged_ranges:
                        merge_info = f" (merged range: {merged_ranges[cell_id]})"
                    
                    markdown_output.append(f"{cell_id}: \"{str(cell_value)}\"{merge_info}  ")

    # Write to markdown file
    output_filename = filename.rsplit('.', 1)[0] + '.md'  # Remove excel extension and add .md
    with open(f'output/{output_filename}', 'w', encoding='utf-8') as f:
        f.write('\n'.join(markdown_output))

if __name__ == "__main__":
    excel_to_markdown("IGEG1688I.xlsx")
